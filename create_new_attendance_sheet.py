import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import argparse
import glob
import re
import sys
import os

def get_weekday_name(date):
    """获取星期名称"""
    weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    return weekdays[date.weekday()]

def parse_punch_times(time_str):
    """解析打卡时间字符串，返回签到和签退时间"""
    if pd.isna(time_str) or time_str == '':
        return None, None
    times = [t.strip() for t in str(time_str).split('\n') if t.strip()]
    if len(times) == 0:
        return None, None
    elif len(times) == 1:
        return times[0], None
    else:
        return times[0], times[-1]

def calculate_work_hours(check_in_time, check_out_time):
    """计算工作时长（小时）"""
    if not check_in_time or not check_out_time:
        return 0.0
    
    def parse_time(time_str):
        """解析时间字符串为小时数"""
        try:
            # 处理 HH:MM 格式
            if ':' in time_str:
                hours, minutes = map(int, time_str.split(':'))
                return hours + minutes / 60.0
            # 处理 HH.MM 格式
            elif '.' in time_str:
                return float(time_str)
            else:
                return float(time_str)
        except:
            return 0.0
    
    check_in_hours = parse_time(check_in_time)
    check_out_hours = parse_time(check_out_time)
    
    if check_in_hours == 0 or check_out_hours == 0:
        return 0.0
    
    # 判断是否跨天
    if check_in_hours > check_out_hours:
        # 跨天情况：签到时间 > 签退时间
        # 计算公式：(24:00 - 签到时间) + 签退时间
        work_hours = (24.0 - check_in_hours) + check_out_hours
    else:
        # 同一天情况：签到时间 < 签退时间
        # 计算公式：签退时间 - 签到时间
        work_hours = check_out_hours - check_in_hours
    
    return max(0.0, work_hours)

def test_time_calculation():
    """测试时间计算逻辑"""
    print("=== 测试时间计算逻辑 ===")
    
    # 测试用例
    test_cases = [
        # (签到时间, 签退时间, 期望结果, 描述)
        ("08:30", "18:30", 10.0, "同一天正常工作时间"),
        ("09:00", "17:00", 8.0, "同一天8小时工作"),
        ("22:00", "06:00", 8.0, "跨天夜班工作"),
        ("23:00", "07:00", 8.0, "跨天夜班工作"),
        ("00:00", "08:00", 8.0, "跨天夜班工作"),
        ("08:00", "08:00", 0.0, "相同时间"),
        ("18:30", "08:30", 14.0, "跨天工作"),
        ("", "18:30", 0.0, "缺少签到时间"),
        ("08:30", "", 0.0, "缺少签退时间"),
        ("", "", 0.0, "缺少所有时间"),
    ]
    
    for check_in, check_out, expected, description in test_cases:
        result = calculate_work_hours(check_in, check_out)
        status = "✓" if abs(result - expected) < 0.01 else "✗"
        print(f"{status} {description}: 签到{check_in} 签退{check_out} -> {result:.2f}小时 (期望{expected:.2f}小时)")
    
    print("=== 测试完成 ===")

def find_source_file(pattern=None):
    """查找数据源文件"""
    if pattern:
        # 如果指定了文件名模式，直接查找
        source_files = glob.glob(pattern)
        if source_files:
            return source_files[0]
    
    # 自动查找符合格式的数据源文件
    pattern = r'考勤表-上下班工时统计表(\d{4})年(\d{1,2})月.*\.xlsx'
    source_files = glob.glob('考勤表-上下班工时统计表*.xlsx')
    
    for file in source_files:
        match = re.match(pattern, file)
        if match:
            return file
    
    return None

def parse_date_from_filename(filename):
    """从文件名解析年月信息"""
    pattern = r'考勤表-上下班工时统计表(\d{4})年(\d{1,2})月.*\.xlsx'
    match = re.match(pattern, filename)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return year, month
    return None, None

def get_days_in_month(year, month):
    """获取指定年月的天数，处理闰年"""
    if month == 2:
        # 2月特殊处理闰年
        if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
            return 29  # 闰年2月29天
        else:
            return 28  # 平年2月28天
    elif month in [4, 6, 9, 11]:
        return 30  # 小月30天
    else:
        return 31  # 大月31天

def create_new_attendance_sheet(source_file=None, output_file=None, year=None, month=None):
    """创建新的考勤统计表"""
    print("正在创建新的考勤统计表...")
    
    # 查找数据源文件
    if not source_file:
        source_file = find_source_file()
    
    if not source_file:
        print("错误：未找到数据源文件！")
        print("请确保当前目录下有符合格式的文件：考勤表-上下班工时统计表YYYY年M月.xlsx")
        return None
    
    print(f"找到数据源文件：{source_file}")
    
    # 解析年月信息
    if not year or not month:
        parsed_year, parsed_month = parse_date_from_filename(source_file)
        if not parsed_year or not parsed_month:
            print("错误：无法从文件名解析年月信息！")
            return None
        year, month = parsed_year, parsed_month
    
    print(f"解析得到：{year}年{month}月")
    
    # 读取原始数据获取员工信息
    try:
        df_original = pd.read_excel(source_file, sheet_name='考勤统计', header=[5, 6])
        df_punch = pd.read_excel(source_file, sheet_name='打卡时间', header=[2, 3])
    except Exception as e:
        print(f"错误：读取Excel文件失败 - {e}")
        return None
    
    # 获取员工列表
    employees = df_original.iloc[:, 0].dropna().tolist()
    positions = df_original.iloc[:, 1].dropna().tolist()
    
    print(f"找到 {len(employees)} 个员工: {employees}")
    print(f"岗位信息: {positions}")
    
    # 计算指定年月的天数
    days_in_month = get_days_in_month(year, month)
    print(f"{year}年{month}月共有{days_in_month}天")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    
    # 设置样式
    header_font = Font(name='微软雅黑', size=12, bold=True)
    date_font = Font(name='微软雅黑', size=10)
    cell_font = Font(name='微软雅黑', size=9)
    
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    date_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建表头
    print("正在创建表头...")
    
    # 第1行：标题
    ws.merge_cells('A1:Z1')
    ws['A1'] = '考勤表-上下班工时统计表'
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = center_alignment
    
    # 第2行：公司信息
    ws.merge_cells('A2:Z2')
    ws['A2'] = '公司名称：武汉福福数通网络科技有限公司'
    ws['A2'].font = cell_font
    ws['A2'].alignment = Alignment(horizontal='left')
    
    # 第3行：空行
    ws['A3'] = ''
    
    # 第4行：时间信息
    ws['A4'] = f'时间段：{year}年{month}月1日-{year}年{month}月{days_in_month}日'
    ws['A4'].font = cell_font
    
    # 第5行：空行
    ws['A5'] = ''
    
    # 第6行：星期标题行
    current_col = 1
    ws.cell(row=6, column=current_col, value='员工姓名')
    ws.cell(row=6, column=current_col+1, value='岗位')
    ws.cell(row=6, column=current_col+2, value='')
    current_col += 3
    
    # 生成指定年月的日期列（每天两列：签到签退列 + 工作时长列）
    start_date = datetime(year, month, 1)
    for day in range(days_in_month):
        current_date = start_date + timedelta(days=day)
        weekday = get_weekday_name(current_date)
        date_str = f"{current_date.month}月{current_date.day}日"
        
        # 每天两列：签到签退列 + 工作时长列
        ws.cell(row=6, column=current_col, value=weekday)
        ws.cell(row=7, column=current_col, value=date_str)
        current_col += 1
        
        # 工作时长列
        ws.cell(row=6, column=current_col, value='工作时长')
        ws.cell(row=7, column=current_col, value='(小时)')
        current_col += 1
    
    # 添加其他列
    ws.cell(row=6, column=current_col, value='累计时长')
    ws.cell(row=6, column=current_col+1, value='休息天数')
    ws.cell(row=6, column=current_col+2, value='备注')
    
    # 应用样式到表头
    for col in range(1, current_col + 3):
        for row in range(6, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
    
    # 创建员工数据行
    print("正在创建员工数据行...")
    current_row = 8
    
    for i, (employee, position) in enumerate(zip(employees, positions)):
        # 签到行
        ws.cell(row=current_row, column=1, value=employee)
        ws.cell(row=current_row, column=2, value=position)
        ws.cell(row=current_row, column=3, value='签到')
        
        # 签退行
        ws.cell(row=current_row+1, column=1, value='')  # 员工姓名留空，后续合并
        ws.cell(row=current_row+1, column=2, value='')  # 岗位留空，后续合并
        ws.cell(row=current_row+1, column=3, value='签退')
        
        # 合并员工姓名单元格（两行合并）
        ws.merge_cells(f'A{current_row}:A{current_row+1}')
        ws.cell(row=current_row, column=1).alignment = center_alignment
        
        # 合并岗位单元格（两行合并）
        ws.merge_cells(f'B{current_row}:B{current_row+1}')
        ws.cell(row=current_row, column=2).alignment = center_alignment
        
        # 应用样式
        for col in range(1, current_col + 3):
            for row in range(current_row, current_row + 2):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = border
                cell.alignment = center_alignment
        
        current_row += 2
    
    # 填充所有员工的数据
    print("正在填充所有员工的打卡数据...")
    
    total_processed = 0
    for employee in employees:
        # 找到员工在打卡时间表中的行
        punch_row = df_punch[df_punch.iloc[:, 0] == employee]
        if not punch_row.empty:
            punch_row = punch_row.iloc[0]
            
            # 找到员工在新表中的行
            employee_new_row = None
            for row in range(8, current_row, 2):
                if ws.cell(row=row, column=1).value == employee:
                    employee_new_row = row
                    break
            
            if employee_new_row:
                employee_processed_count = 0
                
                # 处理指定天数的数据
                total_work_hours = 0  # 累计工作时长
                work_days = 0  # 工作天数
                
                for day in range(1, days_in_month + 1):
                    target_date = datetime(year, month, day)
                    date_str = f"{target_date.month}月{target_date.day}日"
                    
                    # 查找打卡时间
                    punch_col = ('打卡时间', str(day))
                    check_in, check_out = None, None
                    
                    if punch_col in df_punch.columns:
                        punch_time = punch_row[punch_col]
                        check_in, check_out = parse_punch_times(punch_time)
                        if check_in or check_out:
                            employee_processed_count += 1
                    
                    # 找到对应的签到签退列（每天两列：签到签退列 + 工作时长列）
                    for col in range(4, current_col, 2):  # 每次跳2列，因为每天有2列
                        if ws.cell(row=7, column=col).value == date_str:
                            # 签到行
                            if check_in:
                                ws.cell(row=employee_new_row, column=col, value=check_in)
                            else:
                                ws.cell(row=employee_new_row, column=col, value='数据缺失')
                            # 签退行
                            if check_out:
                                ws.cell(row=employee_new_row+1, column=col, value=check_out)
                            else:
                                ws.cell(row=employee_new_row+1, column=col, value='数据缺失')
                            
                            # 计算当天工作时长并填入工作时长列
                            hours_col = col + 1  # 工作时长列在签到签退列的右边
                            
                            if check_in and check_out:
                                daily_hours = calculate_work_hours(check_in, check_out)
                                total_work_hours += daily_hours
                                work_days += 1
                                
                                # 填入工作时长（合并单元格）
                                ws.cell(row=employee_new_row, column=hours_col, value=f"{daily_hours:.2f}")
                                ws.cell(row=employee_new_row+1, column=hours_col, value='')  # 签退行留空
                                
                                print(f"{employee} {date_str}: 签到{check_in} 签退{check_out} 工作时长{daily_hours:.2f}小时")
                            else:
                                # 即使没有工作时长，也要合并单元格并留空
                                ws.cell(row=employee_new_row, column=hours_col, value='')
                                ws.cell(row=employee_new_row+1, column=hours_col, value='')
                            
                            # 始终合并工作时长单元格（两行合并）
                            ws.merge_cells(f'{get_column_letter(hours_col)}{employee_new_row}:{get_column_letter(hours_col)}{employee_new_row+1}')
                            ws.cell(row=employee_new_row, column=hours_col).alignment = center_alignment
                            
                            break
                
                if employee_processed_count > 0:
                    total_processed += employee_processed_count
                    print(f"{employee}: 处理了 {employee_processed_count} 天有打卡记录的日期")
                    
                    # 填入累计时长（合并单元格，像工作时长一样）
                    summary_col = current_col  # 累计时长列
                    ws.cell(row=employee_new_row, column=summary_col, value=f"{total_work_hours:.2f}")
                    ws.cell(row=employee_new_row+1, column=summary_col, value='')  # 签退行留空
                    
                    # 合并累计时长单元格（两行合并）
                    ws.merge_cells(f'{get_column_letter(summary_col)}{employee_new_row}:{get_column_letter(summary_col)}{employee_new_row+1}')
                    ws.cell(row=employee_new_row, column=summary_col).alignment = center_alignment
                    
                    print(f"{employee}: 累计工作时长 {total_work_hours:.2f} 小时，工作天数 {work_days} 天")
    
    print(f"所有员工数据处理完成！总共处理了 {total_processed} 条打卡记录")
    
    # 调整列宽
    for col in range(1, current_col + 3):
        if col <= 3:
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col % 2 == 0:  # 工作时长列
            ws.column_dimensions[get_column_letter(col)].width = 8
        else:  # 签到签退列
            ws.column_dimensions[get_column_letter(col)].width = 10
    
    # 保存文件
    if not output_file:
        output_file = f'{year}年{month}月员工考勤统计表.xlsx'
    
    wb.save(output_file)
    print(f"新考勤统计表已创建: {output_file}")
    print(f"包含 {len(employees)} 个员工，{days_in_month} 天的完整结构")
    print("每天两列：签到签退列 + 工作时长列")
    print("员工姓名和岗位列已合并为两行单元格")
    print("所有员工的打卡数据已导入完成")
    
    return output_file

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='考勤统计表生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python create_new_attendance_sheet.py                    # 自动查找数据源文件
  python create_new_attendance_sheet.py --test             # 运行测试
  python create_new_attendance_sheet.py --input "data.xlsx" # 指定数据源文件
  python create_new_attendance_sheet.py --output "result.xlsx" # 指定输出文件
  python create_new_attendance_sheet.py --year 2025 --month 6 # 指定年月
        """
    )
    
    parser.add_argument('--input', '-i', 
                       help='数据源文件名（可选，默认自动查找）')
    parser.add_argument('--output', '-o', 
                       help='输出文件名（可选）')
    parser.add_argument('--year', '-y', type=int,
                       help='指定年份（可选）')
    parser.add_argument('--month', '-m', type=int,
                       help='指定月份（可选）')
    parser.add_argument('--test', '-t', action='store_true',
                       help='运行测试用例')
    
    args = parser.parse_args()
    
    # 如果只是运行测试
    if args.test:
        test_time_calculation()
        return
    
    # 验证参数
    if args.month and (args.month < 1 or args.month > 12):
        print("错误：月份必须在1-12之间")
        return
    
    if args.year and (args.year < 1900 or args.year > 2100):
        print("错误：年份必须在1900-2100之间")
        return
    
    # 先运行测试
    test_time_calculation()
    print("\n" + "="*50 + "\n")
    
    # 创建考勤表
    result = create_new_attendance_sheet(
        source_file=args.input,
        output_file=args.output,
        year=args.year,
        month=args.month
    )
    
    if result:
        print(f"\n✅ 处理完成！输出文件：{result}")
    else:
        print("\n❌ 处理失败！")
        sys.exit(1)

if __name__ == "__main__":
    main() 